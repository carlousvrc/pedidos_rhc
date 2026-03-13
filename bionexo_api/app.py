#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Servidor web para converter PDFs do Bionexo em Excel.
Uso: python app.py
Acesse: http://localhost:5000
"""

import os
import io
import uuid

import pandas as pd
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from converter_bionexo import process_pdf_buffer

app = Flask(__name__)
CORS(app)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB

# Cache em memoria: token -> lista de registros
_cache = {}

# Colunas removidas no export (internas / pouco uteis para compras)
_DROP_COLS = ["Porcentagem (%)", "Justificativa", "Comentario", "Usuario", "Item"]

_COL_WIDTHS = {
    "Unidade Hospitalar":     30,
    "Pedido de Cotacao":      16,
    "Data Emissao":           13,
    "Fornecedor":             35,
    "Produto":                45,
    "Codigo":                  8,
    "Fabricante":             22,
    "Embalagem":              22,
    "Preco Unitario (R$)":    16,
    "Quantidade":             11,
    "Unidade":                13,
    "Valor Total (R$)":       16,
    "Preco Referencia (R$)":  18,
    "Status":                 16,
}

STATUS_COLORS = {
    "Recebido":          "C6EFCE",  # verde
    "Recebido Parcial":  "FFEB9C",  # amarelo
    "Nao Recebido":      "FFC7CE",  # vermelho
    "Pendente":          "FFFFFF",  # branco
}


def build_excel(data, include_status=True):
    """Gera BytesIO com o Excel formatado. data e uma lista de dicts."""
    if not data:
        return None

    df = pd.DataFrame(data)

    if "Produto" in df.columns:
        df = df[df["Produto"].str.strip().str.len() > 0]

    if df.empty:
        return None

    # Remove colunas internas
    df.drop(columns=[c for c in _DROP_COLS if c in df.columns], inplace=True)

    # Garante coluna Status
    if include_status and "Status" not in df.columns:
        df["Status"] = "Pendente"

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Dados", index=False)
        ws = writer.sheets["Dados"]

        hdr_fill  = PatternFill("solid", fgColor="1F4E79")
        hdr_font  = Font(color="FFFFFF", bold=True, size=10)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = hdr_align
        ws.row_dimensions[1].height = 35

        status_col_idx = None
        for i, col_name in enumerate(df.columns, 1):
            if col_name == "Status":
                status_col_idx = i

        fill_alt = PatternFill("solid", fgColor="DCE6F1")
        for r in range(2, len(df) + 2):
            status_val = ""
            if status_col_idx:
                status_val = str(ws.cell(row=r, column=status_col_idx).value or "")

            for c in range(1, len(df.columns) + 1):
                cell = ws.cell(row=r, column=c)
                col_name = df.columns[c - 1]

                # Cor de linha baseada no status
                if status_col_idx and status_val in STATUS_COLORS:
                    color = STATUS_COLORS[status_val]
                    cell.fill = PatternFill("solid", fgColor=color)
                elif r % 2 == 0:
                    cell.fill = fill_alt

                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if "R$" in col_name:
                    cell.number_format = 'R$ #,##0.00'
                elif "%" in col_name:
                    cell.number_format = '0.00"%"'

                # Negrito na coluna Status
                if col_name == "Status":
                    cell.font = Font(bold=True)

        for i, col_name in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = \
                _COL_WIDTHS.get(col_name, 15)

        ws.freeze_panes    = "A2"
        ws.auto_filter.ref = ws.dimensions

    buf.seek(0)
    return buf


@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})


@app.route("/converter", methods=["POST"])
def converter():
    """Converte PDF e retorna JSON com os dados extraidos + token de cache."""
    if "pdf" not in request.files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400

    file = request.files["pdf"]

    if not file.filename:
        return jsonify({"error": "Nome de arquivo vazio"}), 400

    if not file.filename.lower().endswith(".pdf"):
        return jsonify({"error": "O arquivo deve ser um PDF"}), 400

    pdf_bytes = file.read()

    try:
        data = process_pdf_buffer(pdf_bytes)
    except Exception as e:
        return jsonify({"error": "Erro ao processar PDF: {}".format(str(e))}), 500

    if not data:
        return jsonify({"error": "Nenhum dado extraido. Verifique se o PDF e um relatorio Bionexo valido."}), 422

    token = str(uuid.uuid4())
    _cache[token] = data

    return jsonify({
        "token":    token,
        "count":    len(data),
        "filename": file.filename,
        "data":     data,
    })


@app.route("/export", methods=["POST"])
def export():
    """Gera Excel a partir dos dados JSON (com coluna Status incluida)."""
    body = request.get_json(force=True, silent=True) or {}
    data = body.get("data")
    filename = body.get("filename", "Bionexo_Export.xlsx")

    if not data:
        return jsonify({"error": "Nenhum dado recebido"}), 400

    try:
        buf = build_excel(data, include_status=True)
    except Exception as e:
        return jsonify({"error": "Erro ao gerar Excel: {}".format(str(e))}), 500

    if buf is None:
        return jsonify({"error": "Falha ao gerar planilha Excel"}), 500

    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/consolidar", methods=["POST"])
def consolidar():
    """Consolida multiplos tokens em um unico Excel."""
    body = request.get_json(force=True, silent=True) or {}
    tokens = body.get("tokens", [])

    if len(tokens) < 2:
        return jsonify({"error": "Tokens insuficientes para consolidar"}), 400

    all_data = []
    for token in tokens:
        if token not in _cache:
            return jsonify({"error": "Sessao expirada. Reconverta os arquivos."}), 400
        all_data.extend(_cache[token])

    if not all_data:
        return jsonify({"error": "Nenhum dado disponivel para consolidar"}), 422

    try:
        buf = build_excel(all_data, include_status=True)
    except Exception as e:
        return jsonify({"error": "Erro ao gerar Excel consolidado: {}".format(str(e))}), 500

    if buf is None:
        return jsonify({"error": "Falha ao gerar planilha consolidada"}), 500

    return send_file(
        buf,
        as_attachment=True,
        download_name="Consolidado_Bionexo.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print("=" * 50)
    print("  Conversor Bionexo PDF -> Excel")
    print("  Acesse: http://localhost:{}".format(port))
    print("=" * 50)
    app.run(debug=False, port=port, host="0.0.0.0")
