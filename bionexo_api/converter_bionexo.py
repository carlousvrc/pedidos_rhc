#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Conversor PDF Bionexo -> Excel
================================
Extrai dados de relatorios de cotacao do Bionexo e gera planilha Excel.

Suporta:
  - Relatorios de Materiais Hospitalares
  - Relatorios de Medicamentos (multiplos fornecedores por arquivo)

Uso:
  python converter_bionexo.py                   -> converte todos os PDFs da pasta
  python converter_bionexo.py arquivo.pdf        -> converte arquivo especifico
  python converter_bionexo.py a.pdf b.pdf        -> converte multiplos arquivos

Dependencias instaladas automaticamente:
  pdfplumber, pandas, openpyxl
"""

import sys
import os
import io
import re
from pathlib import Path

import pdfplumber
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ── Funcoes auxiliares ─────────────────────────────────────────────────────────

def parse_brl(text):
    """Extrai valor monetario de um texto misto. Ex: '100MG R$ 1,3900' -> 1.39"""
    if not text:
        return None
    s = str(text)
    # Prioridade: encontra 'R$' seguido de valor numerico
    m = re.search(r"R\$\s*([\d.]+,[\d]+|[\d]+,[\d]+)", s)
    if m:
        val = m.group(1).replace(".", "").replace(",", ".")
        try:
            return float(val)
        except ValueError:
            pass
    # Fallback: ultimo numero com virgula decimal no texto
    matches = re.findall(r"[\d]+(?:\.[\d]{3})*,[\d]+", s)
    if matches:
        val = matches[-1].replace(".", "").replace(",", ".")
        try:
            return float(val)
        except ValueError:
            pass
    return None


def parse_pct(text):
    """Extrai percentual de texto misto. Ex: '-6,08% 03/02/2026' -> -6.08"""
    if not text:
        return None
    s = str(text)
    # Procura padrao: sinal opcional + digitos + virgula/ponto + digitos + %
    m = re.search(r"([+-]?[\d]+[,.][\d]+)\s*%", s)
    if m:
        val = m.group(1).replace(",", ".")
        try:
            return float(val)
        except ValueError:
            pass
    # Fallback: primeiro numero com virgula no texto
    m2 = re.search(r"([+-]?[\d]+[,.][\d]+)", s)
    if m2:
        val = m2.group(1).replace(",", ".")
        try:
            return float(val)
        except ValueError:
            pass
    return None


def parse_qty(text):
    """'288 Frasco'  ->  (288, 'Frasco')"""
    if not text:
        return None, ""
    m = re.match(r"^([\d.]+)\s*(.*)", str(text).strip())
    if m:
        qty_str = m.group(1).replace(".", "")
        unit = m.group(2).strip()
        try:
            return int(qty_str), unit
        except ValueError:
            return qty_str, unit
    return None, str(text).strip()


def clean_ws(text):
    """Remove whitespace excessivo."""
    return re.sub(r"\s+", " ", str(text or "")).strip()


# ── Extracao baseada em posicao de palavras ────────────────────────────────────

# Nomes das colunas que queremos detectar no cabecalho
_HEADER_KEYS = {
    "produto":        "produto",
    "codigo":         "codigo",
    "programacao":    "programacao",
    "fabricante":     "fabricante",
    "embalagem":      "embalagem",
    "fornecedor":     "fornecedor",
    "comentario":     "comentario",
    "unitario":       "preco_unitario",
    "quantidade":     "quantidade",
    "justificativa":  "justificativa",
    "total":          "valor_total",
    "referencia":     "preco_referencia",
    "porcentagem":    "porcentagem",
    "usuario":        "usuario",
}

def _normalize(text):
    """Normaliza para comparacao: minusculo, sem acentos."""
    t = text.lower()
    subs = {
        "a\xe7": "a", "\xe3": "a", "\xe1": "a", "\xe0": "a", "\xe2": "a",
        "\xe9": "e", "\xea": "e", "\xe8": "e",
        "\xed": "i", "\xee": "i",
        "\xf3": "o", "\xf4": "o", "\xf5": "o",
        "\xfa": "u", "\xfc": "u",
        "\xe7": "c", "\xf1": "n",
    }
    for k, v in subs.items():
        t = t.replace(k, v)
    return t


def detect_columns(header_words):
    """
    Dado o conjunto de palavras do cabecalho, retorna lista de
    (col_name, x_start, x_end) ordenada por x_start.
    """
    # Agrupa palavras do cabecalho pelo campo que representam
    buckets = {}  # col_name -> [x0, x1, ...]

    for w in header_words:
        wn = _normalize(w["text"])
        matched = None
        for key, col in _HEADER_KEYS.items():
            if key in wn:
                matched = col
                break
        if matched:
            if matched not in buckets:
                buckets[matched] = {"x0": w["x0"], "x1": w["x1"]}
            else:
                buckets[matched]["x0"] = min(buckets[matched]["x0"], w["x0"])
                buckets[matched]["x1"] = max(buckets[matched]["x1"], w["x1"])

    if not buckets:
        return []

    cols = sorted([(name, info["x0"], info["x1"]) for name, info in buckets.items()],
                  key=lambda c: c[1])

    # Define limites de coluna usando pontos medios entre headers adjacentes.
    # Isso evita sobreposicoes e atribui palavras a coluna mais proxima.
    result = []
    for i, (name, x0, x1) in enumerate(cols):
        x_start = (cols[i - 1][1] + x0) / 2 if i > 0 else x0 - 5
        x_end   = (x0 + cols[i + 1][1]) / 2 if i + 1 < len(cols) else 9999
        result.append((name, x_start, x_end))

    # Adiciona coluna do numero do item (sem header proprio) antes de tudo.
    # Item numbers ficam em x0=35-42; limite maximo de 47px.
    ITEM_NUM_X_MAX = 47
    result.insert(0, ("item_num", 0, ITEM_NUM_X_MAX))
    # Produto deve comecar logo apos item_num (pode ter texto em x0=48+)
    if len(result) > 1:
        name1, _, x_end1 = result[1]
        result[1] = (name1, ITEM_NUM_X_MAX, x_end1)
    return result


def assign_col(x0, columns):
    """Retorna o nome da coluna para um dado x0."""
    for (name, x_start, x_end) in columns:
        if x_start <= x0 <= x_end:
            return name
    return None


def words_to_row(words_in_item, columns):
    """
    Dado um conjunto de palavras pertencentes a um item,
    retorna dicionario {col_name: texto}.
    """
    cells = {}
    for w in words_in_item:
        col = assign_col(w["x0"], columns)
        if col:
            cells.setdefault(col, []).append((w["top"], w["text"]))

    result = {}
    for col, parts in cells.items():
        # Ordena por y (top) depois por x0 para formar o texto
        parts.sort(key=lambda p: p[0])
        result[col] = " ".join(p[1] for p in parts)

    return result


def extract_supplier_name(words_before_header):
    """Tenta extrair o nome do fornecedor das palavras antes do cabecalho."""
    # Procura palavras na area da tabela de fornecedor (geralmente ~10px abaixo
    # de "Fornecedor" no cabecalho da tabela de info)
    cand = []
    for w in words_before_header:
        if w["x0"] < 170 and len(w["text"]) > 3:
            cand.append(w)
    if not cand:
        return ""
    # Pega as primeiras palavras de uma mesma linha
    first_y = min(w["top"] for w in cand)
    first_line = [w for w in cand if abs(w["top"] - first_y) < 8]
    first_line.sort(key=lambda w: w["x0"])
    return " ".join(w["text"] for w in first_line)


def process_page(page, current_supplier, pedido, dt_emissao, last_columns=None):
    """
    Processa uma pagina e retorna (lista_de_registros, fornecedor_atualizado, colunas).
    Em paginas de continuacao (sem cabecalho 'Produto'), reutiliza last_columns.
    """
    words = page.extract_words(x_tolerance=2, y_tolerance=3)
    if not words:
        return [], current_supplier, last_columns

    # Encontra Y do cabecalho da tabela de itens (linha com 'Produto')
    produto_words = [w for w in words if _normalize(w["text"]) == "produto"]

    if produto_words:
        header_y = min(w["top"] for w in produto_words)

        # Palavras do cabecalho (±12px ao redor de header_y)
        header_words = [w for w in words
                        if header_y - 12 <= w["top"] <= header_y + 20]

        # Detecta colunas
        columns = detect_columns(header_words)
        if columns:
            last_columns = columns  # atualiza cache

        # Tenta identificar o fornecedor na secao acima do cabecalho
        forn_area_words = [w for w in words
                           if w["top"] > 80 and w["top"] < header_y - 20
                           and w["x0"] < 170]
        forn_skip = {"fornecedor","faturamento","minimo","prazo","entrega",
                     "validade","proposta","condicoes","pagamento","frete",
                     "observacoes","webservice","av","rua","est","avenida"}
        forn_words = [w for w in forn_area_words
                      if _normalize(w["text"]) not in forn_skip
                      and not re.match(r'^[\d\s\.\-,/@\(\)]+$', w["text"])
                      and len(w["text"]) > 2]
        if forn_words:
            forn_words.sort(key=lambda w: (w["top"], w["x0"]))
            first_y = forn_words[0]["top"]
            first_line = [w for w in forn_words if abs(w["top"] - first_y) < 8]
            first_line.sort(key=lambda w: w["x0"])
            candidate = " ".join(w["text"] for w in first_line)
            if len(candidate) > 5:
                current_supplier = candidate

        # Palavras abaixo do cabecalho (itens)
        item_area = [w for w in words if w["top"] > header_y + 15]

    else:
        # Pagina de continuacao: sem cabecalho, usa colunas do cache
        columns = last_columns
        item_area = words  # todos os itens comecam do topo da pagina

    if not columns:
        return [], current_supplier, last_columns

    # Encontra posicoes y dos numeros de item (numero puro na coluna mais a esquerda)
    # Usa um limite x0 conservador: no maximo 60% da largura da coluna item_num
    item_col_start = columns[0][1]
    item_col_end   = columns[0][2]  # coluna "item_num"
    # Produto comeca em columns[1][1]; item numbers ficam bem a esquerda disso
    item_x_max = columns[1][1] - 5 if len(columns) > 1 else item_col_end

    # Coleta candidatos e remove duplicatas por Y (fica so o mais a esquerda)
    candidates_by_y = {}
    for w in item_area:
        if (re.match(r"^\d{1,3}$", w["text"])
                and item_col_start <= w["x0"] <= item_x_max):
            y = round(w["top"], 1)
            if y not in candidates_by_y or w["x0"] < candidates_by_y[y][1]:
                candidates_by_y[y] = (w["text"], w["x0"])

    item_starts = [(y, txt) for y, (txt, _) in candidates_by_y.items()]

    if not item_starts:
        return [], current_supplier, last_columns

    item_starts.sort(key=lambda x: x[0])

    # Para cada item, coleta palavras usando pontos medios entre itens consecutivos.
    # Isso garante que texto multi-linha (antes do numero do item) seja capturado.
    records = []
    max_y = max(w["top"] for w in item_area)

    for idx, (y_start, item_num) in enumerate(item_starts):
        # Inicio: ponto medio entre item anterior e este (ou inicio da area de itens)
        if idx == 0:
            # Primeira pagina com header: vai do inicio da area de itens
            # Continuacao: vai 15px antes do numero do item (sem pegar header da pagina)
            if produto_words:
                y_collect_start = min(w["top"] for w in item_area) if item_area else y_start
            else:
                y_collect_start = max(
                    min(w["top"] for w in item_area) if item_area else 0,
                    y_start - 15
                )
        else:
            prev_y = item_starts[idx - 1][0]
            y_collect_start = (prev_y + y_start) / 2

        # Fim: ponto medio entre este item e o proximo
        if idx + 1 < len(item_starts):
            next_y = item_starts[idx + 1][0]
            y_collect_end = (y_start + next_y) / 2
        else:
            y_collect_end = max_y + 10

        item_words = [w for w in item_area
                      if y_collect_start <= w["top"] <= y_collect_end]

        row = words_to_row(item_words, columns)

        if not row.get("produto"):
            continue

        qty_raw = row.get("quantidade", "")
        qty, unit = parse_qty(qty_raw)

        forn_in_row = clean_ws(row.get("fornecedor", ""))

        records.append({
            "Pedido de Cotacao":   pedido,
            "Data Emissao":        dt_emissao,
            "Fornecedor":          forn_in_row or current_supplier,
            "Item":                item_num,
            "Produto":             clean_ws(row.get("produto", "")),
            "Codigo":              clean_ws(row.get("codigo", "")),
            "Fabricante":          clean_ws(row.get("fabricante", "")),
            "Embalagem":           clean_ws(row.get("embalagem", "")),
            "Comentario":          clean_ws(row.get("comentario", "")),
            "Preco Unitario (R$)": parse_brl(row.get("preco_unitario", "")),
            "Quantidade":          qty,
            "Unidade":             unit,
            "Justificativa":       clean_ws(row.get("justificativa", "")),
            "Valor Total (R$)":    parse_brl(row.get("valor_total", "")),
            "Preco Referencia (R$)": parse_brl(row.get("preco_referencia", "")),
            "Porcentagem (%)":     parse_pct(row.get("porcentagem", "")),
            "Usuario":             clean_ws(row.get("usuario", "")),
        })

    return records, current_supplier, last_columns


def process_pdf(pdf_path):
    """Processa todo o PDF e retorna lista de registros."""
    pdf_path = Path(pdf_path)
    print("\nProcessando: {}".format(pdf_path.name))

    all_records = []
    pedido       = ""
    dt_emissao   = ""
    supplier     = ""
    last_columns = None

    with pdfplumber.open(pdf_path) as pdf:
        print("  Paginas: {}".format(len(pdf.pages)))

        for pn, page in enumerate(pdf.pages):
            txt = page.extract_text() or ""

            if not pedido:
                m = re.search(r"Pedido de Cota[cç][aã]o\s*[:\s]+(\d+)", txt)
                if m:
                    pedido = m.group(1)

            if not dt_emissao:
                m = re.search(r"Relat[oó]rio emitido em\s+(\d{2}/\d{2}/\d{4})", txt)
                if m:
                    dt_emissao = m.group(1)

            recs, supplier, last_columns = process_page(
                page, supplier, pedido, dt_emissao, last_columns)
            all_records.extend(recs)

    print("  Registros extraidos: {}".format(len(all_records)))
    return all_records


# ── Exportacao para Excel ──────────────────────────────────────────────────────

_COL_WIDTHS = {
    "Unidade Hospitalar":     30,
    "Pedido de Cotacao":      16,
    "Data Emissao":           13,
    "Fornecedor":             35,
    "Item":                    7,
    "Produto":                45,
    "Codigo":                  8,
    "Fabricante":             22,
    "Embalagem":              22,
    "Comentario":             28,
    "Preco Unitario (R$)":    16,
    "Quantidade":             11,
    "Unidade":                13,
    "Justificativa":          35,
    "Valor Total (R$)":       16,
    "Preco Referencia (R$)":  18,
    "Porcentagem (%)":        13,
    "Usuario":                16,
}


def save_excel(data, output_path):
    if not data:
        print("  AVISO: Sem dados para salvar em {}".format(Path(output_path).name))
        return False

    df = pd.DataFrame(data)

    if "Produto" in df.columns:
        df = df[df["Produto"].str.strip().str.len() > 0]

    if df.empty:
        print("  AVISO: DataFrame vazio apos filtragem.")
        return False

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Dados", index=False)
        ws = writer.sheets["Dados"]

        # Cabecalho
        hdr_fill  = PatternFill("solid", fgColor="1F4E79")
        hdr_font  = Font(color="FFFFFF", bold=True, size=10)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = hdr_align
        ws.row_dimensions[1].height = 35

        # Linhas alternadas + formato numerico
        fill_alt = PatternFill("solid", fgColor="DCE6F1")
        for r in range(2, len(df) + 2):
            for c in range(1, len(df.columns) + 1):
                cell = ws.cell(row=r, column=c)
                if r % 2 == 0:
                    cell.fill = fill_alt
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                col_name = df.columns[c - 1]
                if "R$" in col_name:
                    cell.number_format = 'R$ #,##0.00'
                elif "%" in col_name:
                    cell.number_format = '0.00"%"'

        # Largura das colunas
        for i, col_name in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = \
                _COL_WIDTHS.get(col_name, 15)

        ws.freeze_panes   = "A2"
        ws.auto_filter.ref = ws.dimensions

    print("  OK - Salvo: {}  ({} registros)".format(
        Path(output_path).name, len(df)))
    return True


# ── API para uso web (BytesIO) ─────────────────────────────────────────────────

def process_pdf_buffer(pdf_bytes):
    """Processa PDF a partir de bytes e retorna lista de registros."""
    all_records = []
    pedido       = ""
    dt_emissao   = ""
    supplier     = ""
    last_columns = None

    hospital = ""

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for pn, page in enumerate(pdf.pages):
            txt = page.extract_text() or ""

            if not pedido:
                m = re.search(r"Pedido de Cota[cç][aã]o\s*[:\s]+(\d+)", txt)
                if m:
                    pedido = m.group(1)

            if not dt_emissao:
                m = re.search(r"Relat[oó]rio emitido em\s+(\d{2}/\d{2}/\d{4})", txt)
                if m:
                    dt_emissao = m.group(1)

            if not hospital:
                lines = txt.split("\n")
                for i, line in enumerate(lines):
                    if line.strip() == "Comprador" and i + 1 < len(lines):
                        raw = lines[i + 1].strip()
                        hospital = raw.split("(")[0].split(",")[0].strip()
                        break

            recs, supplier, last_columns = process_page(
                page, supplier, pedido, dt_emissao, last_columns)
            all_records.extend(recs)

    if hospital:
        for i, rec in enumerate(all_records):
            all_records[i] = {"Unidade Hospitalar": hospital, **rec}

    return all_records


def save_excel_buffer(data):
    """Salva dados em Excel e retorna BytesIO (None se sem dados)."""
    if not data:
        return None

    df = pd.DataFrame(data)

    if "Produto" in df.columns:
        df = df[df["Produto"].str.strip().str.len() > 0]

    if df.empty:
        return None

    drop_cols = ["Porcentagem (%)", "Justificativa", "Comentario", "Usuario", "Item"]
    df.drop(columns=[c for c in drop_cols if c in df.columns], inplace=True)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Dados", index=False)
        ws = writer.sheets["Dados"]

        hdr_fill  = PatternFill("solid", fgColor="1F4E79")
        hdr_font  = Font(color="FFFFFF", bold=True, size=10)
        hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = hdr_align
        ws.row_dimensions[1].height = 35

        fill_alt = PatternFill("solid", fgColor="DCE6F1")
        for r in range(2, len(df) + 2):
            for c in range(1, len(df.columns) + 1):
                cell = ws.cell(row=r, column=c)
                if r % 2 == 0:
                    cell.fill = fill_alt
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                col_name = df.columns[c - 1]
                if "R$" in col_name:
                    cell.number_format = 'R$ #,##0.00'
                elif "%" in col_name:
                    cell.number_format = '0.00"%"'

        for i, col_name in enumerate(df.columns, 1):
            ws.column_dimensions[get_column_letter(i)].width = \
                _COL_WIDTHS.get(col_name, 15)

        ws.freeze_panes    = "A2"
        ws.auto_filter.ref = ws.dimensions

    buf.seek(0)
    return buf


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) > 1:
        pdf_files = [Path(f) for f in sys.argv[1:] if f.lower().endswith(".pdf")]
    else:
        script_dir = Path(__file__).parent
        pdf_files  = sorted(script_dir.glob("*.pdf"))

    if not pdf_files:
        print("Nenhum arquivo PDF encontrado.")
        print("Uso: python converter_bionexo.py [arquivo.pdf ...]")
        sys.exit(1)

    print("=" * 60)
    print("  Conversor PDF Bionexo -> Excel")
    print("=" * 60)
    print("Arquivos: {}".format(len(pdf_files)))
    for f in pdf_files:
        print("  - {}".format(f.name))

    results = []
    for pdf_file in pdf_files:
        if not pdf_file.exists():
            print("\nERRO: arquivo nao encontrado - {}".format(pdf_file))
            continue
        data = process_pdf(pdf_file)
        if data:
            out = pdf_file.with_suffix(".xlsx")
            ok  = save_excel(data, out)
            if ok:
                results.append((pdf_file.name, out.name, len(data)))
        else:
            print("  AVISO: Nenhum dado extraido de {}".format(pdf_file.name))

    print("\n" + "=" * 60)
    print("RESUMO")
    print("=" * 60)
    if results:
        for src, dst, n in results:
            print("  {} -> {} ({} itens)".format(src, dst, n))
        print("\nConversao concluida!")
    else:
        print("Nenhum arquivo convertido.")


if __name__ == "__main__":
    main()
